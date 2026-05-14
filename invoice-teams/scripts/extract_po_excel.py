"""
注文書Excel から請求書転記用データを抽出するスクリプト。
Usage: python extract_po_excel.py <注文書Excelパス>
"""
import sys
import json
import openpyxl


def extract_from_excel(po_path):
    wb = openpyxl.load_workbook(po_path, data_only=True)
    # 「注文書」シートを優先、なければ最初のシートを使用
    ws = wb['注文書'] if '注文書' in wb.sheetnames else wb.active

    data = {
        '宛名会社': '',
        '担当者': '',
        '件名': '',
        '明細': [],
    }

    for row in ws.iter_rows(min_row=1, max_row=50, values_only=True):
        non_none = {i: v for i, v in enumerate(row) if v is not None}
        if not non_none:
            continue

        col0 = str(non_none.get(0, ''))

        # 発注元会社名：「セールス番号」行の列44
        if 'セールス番号' in col0:
            data['宛名会社'] = str(non_none.get(44, '')).strip()

        # 担当者名：「担当者」を含むラベルがある行の列44
        # ラベルは注文書によって異なる（例：「ジェイアール東日本企画担当者」「〇〇担当者」）
        for col_idx, val in non_none.items():
            if isinstance(val, str) and '担当者' in val and col_idx != 44:
                if 44 in non_none:
                    data['担当者'] = str(non_none[44]).replace('\u3000', ' ').strip()

        # 注文件名：「注文件名」行の列14
        if col0 == '注文件名':
            data['件名'] = str(non_none.get(14, '')).strip()

        # 明細：NO列が数字文字列かつ内容（列2）がある行
        no_val = non_none.get(0)
        if isinstance(no_val, (str, int)):
            no_str = str(no_val).strip()
            if no_str.isdigit() and non_none.get(2):
                qty_raw = str(non_none.get(41, '1式'))
                qty_num = ''.join(filter(str.isdigit, qty_raw)) or '1'
                data['明細'].append({
                    '内容': '■' + str(non_none.get(2, '')).strip(),
                    '数量': int(qty_num),
                    '単価': non_none.get(46, 0),
                })

    return data


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python extract_po_excel.py <注文書Excelパス>")
        sys.exit(1)

    result = extract_from_excel(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2))
